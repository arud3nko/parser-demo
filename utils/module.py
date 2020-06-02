from bs4 import BeautifulSoup
import shutil

import openpyxl
from openpyxl.utils import get_column_letter
import csv
from selenium.webdriver.common.keys import Keys
import time
import random
import os

from datetime import date, timedelta
from options import *

import utils

today = date.today()
pages = []
adverts = []
FILENAME = 'total.csv'


class Advert:
    def __init__(self, title, price, addr, phone):
        self.title = title
        self.price = price
        self.addr = addr
        self.phone = phone


def inner(input_addr, input_mail):
    
    driver = webdriver.Chrome(executable_path=path_driver,
                              chrome_options=options)

    j = driver.get(input_addr)

    title_count = int(driver.find_element_by_class_name("page-title-count-1oJOc").text)

    h = int(title_count / 50) + 1

    index = title_count  # total elements
    counter = 1  # visual counter
    list_count = 0  # additional counter

    if h >= 2:
        title_count = 50  # more than 1 page

    yield 'Страниц для парсинга: ' + str(h) + ' | Получаем код страниц...\n'

    for i in range(h):
        if "?" in input_addr:
            driver.get(input_addr + "&p=" + str(i + 1))
        else:
            driver.get(input_addr + "?p=" + str(i + 1))

        show_phone_buttons = driver.find_elements_by_class_name("js-item-extended-contacts")

        for i in range(title_count):

            show_phone_buttons[i].send_keys("\n")
            time.sleep(random.uniform(1, 3))

            webdriver.ActionChains(driver).send_keys(Keys.ESCAPE).perform()

            time_left = timedelta(seconds=2*(index-counter))
            yield str(counter) + '/' + str(index) + '  |  ' + 'Времени осталось: ' + str(time_left) + '\n'
            counter += 1

        list_count += title_count

        if (index - list_count) < 50:
            title_count = index - list_count

        pages.append(driver.page_source)

    driver.quit()

    yield 'Код страницы получен! Выбираем данные...\n'

    for page in pages:

        soup = BeautifulSoup(page, "html.parser")

        category = soup.find('div', class_='js-catalog_serp') \
            .find_all('div',
                      class_='snippet-horizontal item item_table clearfix js-catalog-item-enum item-with-contact '
                             'js-item-extended') 

        category_elements = category[:index]

        for i in category_elements:
            try:
                title = i.find('h3', class_='snippet-title') \
                    .find('a', class_='snippet-link').text.replace('\n', '')
                price = i.find('span', class_='snippet-price').text.replace('\n', '')
                address = i.find('span', class_='item-address__string')
                if address is None:
                    address = i.find('span', class_='item-address-georeferences-item__content')
                    if address is None:
                        address = i.find('div', class_='data')
                address = address.text.replace('\n', '')
                phone_encrypted_image = i.find('div', class_='item__line') \
                    .find('div', class_='item_table-wrapper') \
                    .find('div', class_='item_table-aside') \
                    .find('div',
                          class_='js-item_table-extended-contacts snippet-contacts item_table-extended-contacts '
                                 'item_table-extended-contacts_transparent is-loading') 
                if phone_encrypted_image is None:
                    phone = 'Номер телефона не указан'
                else:
                    phone = utils.decode.decode(phone_encrypted_image)
                adverts.append(Advert(title, price, address, phone))
            except Exception as e:
                print(e)
                pass

    with open(FILENAME, "w", newline="", encoding='utf-8-sig', errors='ignore') as file:
        columns = ['Наименование', 'Цена', 'Адрес', 'Телефон']
        writer = csv.DictWriter(file, dialect='excel', delimiter=',', fieldnames=columns, quotechar='"',
                                quoting=csv.QUOTE_ALL)
        writer.writeheader()
        f = 0

        for i in range(len(adverts)):
            array = {
                "Наименование": adverts[f].title,
                "Цена": adverts[f].price,
                "Адрес": adverts[f].addr,
                "Телефон": adverts[f].phone
            }
            try:
                writer.writerow(array)
                f = f + 1
            except Exception as e:
                print(e)
                pass

    CSVFile = FILENAME
    XLSFile = "total"
    directory = "total_" + str(today)
    main_dic = './'
    dirsv = './' + directory + '/total.xlsx'

    if os.path.exists(main_dic + directory):
        shutil.rmtree(main_dic + directory)
        time.sleep(1)
        os.makedirs(main_dic + directory)
    else:
        os.makedirs(main_dic + directory)

    yield 'Осталось совсем чуть-чуть... \n'

    def csv_d(CSVFile, XLSFile, directory):
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(CSVFile, 'rt', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for r, row in enumerate(reader, start=1):
                for c, val in enumerate(row, start=1):
                    i = get_column_letter(c)
                    ws.column_dimensions[i].width = 40
                    ws.cell(row=r, column=c).value = val
        wb.save(main_dic + directory + '/' + XLSFile + '.xlsx')

    adverts.clear()
    pages.clear()
    csv_d(CSVFile, XLSFile, directory)
    os.remove(CSVFile)

    utils.mail.send_total(input_addr, index, dirsv, input_mail)

    time.sleep(2)

    shutil.rmtree(main_dic+directory)
